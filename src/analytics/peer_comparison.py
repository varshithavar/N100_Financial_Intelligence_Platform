import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# ---------------------------------------------------
# Paths
# ---------------------------------------------------

BASE_DIR = Path(__file__).resolve().parents[2]
DB_PATH = BASE_DIR / "database" / "nifty100.db"

OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

OUTPUT_FILE = OUTPUT_DIR / "peer_comparison.xlsx"

# ---------------------------------------------------
# Database Connection
# ---------------------------------------------------

conn = sqlite3.connect(DB_PATH)

companies = pd.read_sql("""
SELECT
    company_id,
    company_name,
    sector
FROM companies
""", conn)

financial = pd.read_sql("""
SELECT
    company_id,
    financial_year,
    net_profit_margin_pct,
    return_on_equity_pct,
    debt_to_equity,
    asset_turnover,
    free_cash_flow
FROM financial_ratios
""", conn)

peer = pd.read_sql("""
SELECT
    company_id,
    roe_percentile,
    npm_percentile,
    debt_equity_percentile,
    asset_turnover_percentile,
    free_cash_flow_percentile
FROM peer_percentiles
""", conn)

conn.close()

# ---------------------------------------------------
# Merge Tables
# ---------------------------------------------------

df = companies.merge(financial, on="company_id")
df = df.merge(peer, on="company_id")

# ---------------------------------------------------
# Workbook
# ---------------------------------------------------

wb = Workbook()
wb.remove(wb.active)

green_fill = PatternFill(
    fill_type="solid",
    start_color="90EE90",
    end_color="90EE90"
)

yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFF59D",
    end_color="FFF59D"
)

red_fill = PatternFill(
    fill_type="solid",
    start_color="FF9999",
    end_color="FF9999"
)

header_fill = PatternFill(
    fill_type="solid",
    start_color="4F81BD",
    end_color="4F81BD"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

# ---------------------------------------------------
# Create one sheet per sector
# ---------------------------------------------------

for sector in sorted(df["sector"].unique()):

    ws = wb.create_sheet(title=f"Sector_{sector}")

    sector_df = df[df["sector"] == sector].copy()

    headers = list(sector_df.columns)

    # Header Row
    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Data Rows
    for r, row in enumerate(
        sector_df.itertuples(index=False),
        start=2
    ):

        for c, value in enumerate(row, start=1):

            cell = ws.cell(row=r, column=c)
            cell.value = value

    # Percentile columns
    percentile_cols = [
        "roe_percentile",
        "npm_percentile",
        "debt_equity_percentile",
        "asset_turnover_percentile",
        "free_cash_flow_percentile"
    ]

    for col_name in percentile_cols:

        col_index = headers.index(col_name) + 1

        for row in range(2, len(sector_df) + 2):

            cell = ws.cell(row=row, column=col_index)

            if cell.value is None:
                continue

            if cell.value >= 75:
                cell.fill = green_fill

            elif cell.value <= 25:
                cell.fill = red_fill

            else:
                cell.fill = yellow_fill

    # Median Row
    median_row = len(sector_df) + 3

    ws.cell(
        row=median_row,
        column=1
    ).value = "Median"

    for idx, column in enumerate(headers[1:], start=2):

        if pd.api.types.is_numeric_dtype(sector_df[column]):

            ws.cell(
                row=median_row,
                column=idx
            ).value = float(sector_df[column].median())

# ---------------------------------------------------
# Save Workbook
# ---------------------------------------------------

wb.save(OUTPUT_FILE)

print("=" * 50)
print("Peer Comparison Report Generated")
print("=" * 50)
print("Saved to:")
print(OUTPUT_FILE)