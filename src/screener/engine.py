import sqlite3
import pandas as pd
import yaml


DB_PATH = "database/nifty100.db"
CONFIG_PATH = "config/screener_config.yaml"


# -----------------------------
# Load YAML Configuration
# -----------------------------
def load_config():

    with open(CONFIG_PATH, "r") as file:
        config = yaml.safe_load(file)

    return config["filters"]



# -----------------------------
# Load Financial Data
# -----------------------------
def load_financial_data():

    conn = sqlite3.connect(DB_PATH)

    query = """
    SELECT
        r.company_id,
        r.roe,
        r.pe_ratio,
        r.debt_equity,

        p.revenue,
        p.net_profit,
        p.eps,

        c.operating_cf,
        c.investing_cf,

        b.total_assets

    FROM ratios r

    JOIN profit_loss p
    ON r.company_id = p.company_id

    JOIN cash_flow c
    ON r.company_id = c.company_id

    JOIN balance_sheet b
    ON r.company_id = b.company_id
    """

    df = pd.read_sql(query, conn)

    conn.close()


    # Calculate additional metrics

    df["free_cash_flow"] = (
        df["operating_cf"]
        -
        df["investing_cf"]
    )


    df["asset_turnover"] = (
        df["revenue"]
        /
        df["total_assets"]
    )


    return df



# -----------------------------
# Dynamic Filter Engine
# -----------------------------
def apply_filters(df, filters):

    result = df.copy()


    if "roe_min" in filters:
        result = result[
            result["roe"] >= filters["roe_min"]
        ]


    if "debt_equity_max" in filters:
        result = result[
            result["debt_equity"] <= filters["debt_equity_max"]
        ]


    if "fcf_min" in filters:
        result = result[
            result["free_cash_flow"] >= filters["fcf_min"]
        ]


    if "sales_min" in filters:
        result = result[
            result["revenue"] >= filters["sales_min"]
        ]


    return result



# -----------------------------
# Sprint 3 Preset Screeners
# -----------------------------

def quality_compounder(df):

    return df[
        (df["roe"] > 15)
        &
        (df["debt_equity"] < 1)
        &
        (df["free_cash_flow"] > 0)
    ]



def value_pick(df):

    return df[
        (df["pe_ratio"] < 20)
        &
        (df["debt_equity"] < 2)
    ]



def growth_accelerator(df):

    return df[
        (df["debt_equity"] < 2)
    ]



def dividend_champion(df):

    return df[
        (df["free_cash_flow"] > 0)
    ]



def debt_free_blue_chip(df):

    return df[
        (df["debt_equity"] == 0)
        &
        (df["roe"] > 12)
        &
        (df["revenue"] > 5000)
    ]



def turnaround_watch(df):

    return df[
        (df["free_cash_flow"] > 0)
    ]



# -----------------------------
# Composite Quality Score
# -----------------------------
def calculate_score(df):

    df = df.copy()

    if len(df) == 0:
        df["composite_quality_score"] = []
        return df


    df["composite_quality_score"] = (

        df["roe"].rank(pct=True) * 40

        +

        df["free_cash_flow"].rank(pct=True) * 30

        +

        df["revenue"].rank(pct=True) * 30

    )


    return df



# -----------------------------
# Excel Export with Colour-Coding
# -----------------------------
def export_to_excel(results_dict, filters, output_path="output/screener_output.xlsx"):
    import openpyxl
    from openpyxl.styles import PatternFill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for preset_name, df in results_dict.items():
            sheet_name = preset_name[:31]  # Excel sheet name limit
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Re-open with openpyxl to apply colour-coding
    wb = openpyxl.load_workbook(output_path)

    for preset_name in results_dict:
        sheet_name = preset_name[:31]
        ws = wb[sheet_name]
        df = results_dict[preset_name]
        headers = [cell.value for cell in ws[1]]

        for row_idx in range(2, ws.max_row + 1):
            if "roe_min" in filters and "roe" in headers:
                col = headers.index("roe") + 1
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = green_fill if cell.value is not None and cell.value >= filters["roe_min"] else red_fill

            if "debt_equity_max" in filters and "debt_equity" in headers:
                col = headers.index("debt_equity") + 1
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = green_fill if cell.value is not None and cell.value <= filters["debt_equity_max"] else red_fill

    wb.save(output_path)
    print(f"\nSaved: {output_path}")



# -----------------------------
# Main Execution
# -----------------------------
if __name__ == "__main__":

    data = load_financial_data()
    print("Total companies loaded:", len(data))

    filters = load_config()

    presets = {
        "Quality Compounder": quality_compounder,
        "Value Pick": value_pick,
        "Growth Accelerator": growth_accelerator,
        "Dividend Champion": dividend_champion,
        "Debt-Free Blue Chip": debt_free_blue_chip,
        "Turnaround Watch": turnaround_watch
    }

    results = {}

    for name, function in presets.items():
        result = function(data)
        result = calculate_score(result)
        result = result.sort_values("composite_quality_score", ascending=False)

        print("\n========================")
        print(name)
        print("========================")
        print(result)
        print("Companies found:", len(result))

        results[name] = result

    export_to_excel(results, filters)