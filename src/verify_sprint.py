"""
Sprint 3 Verification Script
Checks deliverables and exit criteria for Epics 03 & 04 (Screener + Peer Engine)

Run from your project root:
    python verify_sprint.py
"""

import os
import sys
import subprocess

PASS = "PASS"
FAIL = "FAIL"
results = []


def check(label, condition, detail=""):
    status = PASS if condition else FAIL
    results.append((label, status, detail))
    print(f"[{status}] {label}" + (f" -- {detail}" if detail else ""))


def section(title):
    print("\n" + "=" * 60)
    print(title)
    print("=" * 60)


# ---------------------------------------------------------------
section("1. Required files exist")
# ---------------------------------------------------------------
required_files = [
    "output/screener_output.xlsx",
    "output/peer_comparison.xlsx",
    "config/screener_config.yaml",
    "src/screener/engine.py",
    "src/analytics/peer.py",
]
for f in required_files:
    check(f"File exists: {f}", os.path.isfile(f))

radar_dir = "reports/radar_charts"
radar_count = 0
if os.path.isdir(radar_dir):
    radar_count = len([f for f in os.listdir(radar_dir) if f.lower().endswith(".png")])
check(f"Radar charts folder has PNGs", radar_count > 0, f"{radar_count} PNG files found")


# ---------------------------------------------------------------
section("2. screener_output.xlsx -- 6 preset sheets, 5-50 rows each")
# ---------------------------------------------------------------
try:
    import openpyxl

    if os.path.isfile("output/screener_output.xlsx"):
        wb = openpyxl.load_workbook("output/screener_output.xlsx", data_only=True)
        sheets = wb.sheetnames
        check("Exactly 6 sheets in screener_output.xlsx", len(sheets) == 6, f"found {len(sheets)}: {sheets}")

        for sheet in sheets:
            ws = wb[sheet]
            row_count = ws.max_row - 1  # minus header
            in_range = 5 <= row_count <= 50
            check(f"Preset '{sheet}' returns 5-50 companies", in_range, f"{row_count} rows")
    else:
        check("screener_output.xlsx checks skipped", False, "file not found")
except ImportError:
    print("openpyxl not installed -- run: pip install openpyxl")
except Exception as e:
    check("screener_output.xlsx readable", False, str(e))


# ---------------------------------------------------------------
section("3. peer_comparison.xlsx -- exactly 11 sheets")
# ---------------------------------------------------------------
try:
    if os.path.isfile("output/peer_comparison.xlsx"):
        wb2 = openpyxl.load_workbook("output/peer_comparison.xlsx", data_only=True)
        sheets2 = wb2.sheetnames
        check("Exactly 11 sheets in peer_comparison.xlsx", len(sheets2) == 11, f"found {len(sheets2)}: {sheets2}")
    else:
        check("peer_comparison.xlsx checks skipped", False, "file not found")
except Exception as e:
    check("peer_comparison.xlsx readable", False, str(e))


# ---------------------------------------------------------------
section("4. Spot-check: Quality Compounder preset (ROE > 15%, D/E < 1)")
# ---------------------------------------------------------------
try:
    import pandas as pd

    if os.path.isfile("output/screener_output.xlsx"):
        df = pd.read_excel("output/screener_output.xlsx", sheet_name="Quality Compounder")
        roe_col = next((c for c in df.columns if "roe" in c.lower()), None)
        de_col = next((c for c in df.columns if c.lower() in ("d/e", "de", "debt_to_equity", "debt/equity")), None)

        if roe_col and de_col:
            bad_rows = df[(df[roe_col] <= 15) | (df[de_col] >= 1)]
            check(
                "All Quality Compounder rows meet ROE>15% and D/E<1",
                len(bad_rows) == 0,
                f"{len(bad_rows)} violating rows" if len(bad_rows) else "verified",
            )
        else:
            check("Quality Compounder column check", False, f"couldn't find ROE/D-E columns: {list(df.columns)}")
    else:
        check("Quality Compounder spot-check skipped", False, "file not found")
except ImportError:
    print("pandas not installed -- run: pip install pandas")
except Exception as e:
    check("Quality Compounder spot-check", False, str(e))


# ---------------------------------------------------------------
section("5. Spot-check: IT Services peer group -- highest ROE = highest percentile")
# ---------------------------------------------------------------
try:
    if os.path.isfile("output/peer_comparison.xlsx"):
        wb3 = openpyxl.load_workbook("output/peer_comparison.xlsx", data_only=True)
        it_sheet_name = next((s for s in wb3.sheetnames if "it" in s.lower()), None)
        if it_sheet_name:
            df_it = pd.read_excel("output/peer_comparison.xlsx", sheet_name=it_sheet_name)
            roe_col = next((c for c in df_it.columns if "roe" in c.lower() and "percentile" not in c.lower()), None)
            roe_pct_col = next((c for c in df_it.columns if "roe" in c.lower() and "percentile" in c.lower()), None)
            if roe_col and roe_pct_col:
                top_roe_row = df_it.loc[df_it[roe_col].idxmax()]
                top_pct_row = df_it.loc[df_it[roe_pct_col].idxmax()]
                same_company = top_roe_row.get("company_id") == top_pct_row.get("company_id")
                check("Highest ROE company also has highest ROE percentile (IT Services)", same_company)
            else:
                check("IT Services ROE percentile check", False, f"columns not found: {list(df_it.columns)}")
        else:
            check("IT Services sheet found", False, f"sheets available: {wb3.sheetnames}")
    else:
        check("Peer percentile spot-check skipped", False, "file not found")
except Exception as e:
    check("IT Services peer spot-check", False, str(e))


# ---------------------------------------------------------------
section("6. Data Quality unit tests -- expect 14 passed, 0 failed")
# ---------------------------------------------------------------
if os.path.isdir("tests"):
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pytest", "tests/", "-v", "--tb=short"],
            capture_output=True,
            text=True,
        )
        output = result.stdout + result.stderr
        print(output[-3000:])
        passed = result.returncode == 0
        check("pytest run completed with 0 failures", passed)
        if "14 passed" not in output:
            check("Exactly 14 DQ tests ran", False, "test count doesn't mention '14 passed' -- check manually")
    except Exception as e:
        check("Running pytest", False, str(e))
else:
    check("tests/ folder exists", False)


# ---------------------------------------------------------------
section("SUMMARY")
# ---------------------------------------------------------------
total = len(results)
passed = sum(1 for _, s, _ in results if s == PASS)
failed = total - passed
print(f"\n{passed}/{total} checks passed, {failed} failed.\n")

if failed:
    print("Failed checks:")
    for label, status, detail in results:
        if status == FAIL:
            print(f"  - {label}" + (f" ({detail})" if detail else ""))